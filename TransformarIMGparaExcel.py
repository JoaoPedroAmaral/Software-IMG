
from tkinter import filedialog
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import numpy as np
import cv2
from colormath.color_objects import sRGBColor, CMYKColor
from colormath.color_conversions import convert_color

def rgb_to_cmyk(rgb):
        # Convertendo o valor RGB para sRGBColor
        rgb_color = sRGBColor(rgb[0] / 255, rgb[1] / 255, rgb[2] / 255)
        # Convertendo de sRGB para CMYK
        cmyk_color = convert_color(rgb_color, CMYKColor)
        # Retornando os valores CMYK normalizados entre 0 e 100
        return [cmyk_color.cmyk_c * 100, cmyk_color.cmyk_m * 100, cmyk_color.cmyk_y * 100, cmyk_color.cmyk_k * 100]

class IMGExcel:
    def __init__(self, root):
        self.root = root
        

    def image_to_excel_cmyk(self,image_path, excel_path):
        # Leitura da imagem
        img = cv2.imread(image_path)
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Criando um novo arquivo Excel
        wb = Workbook()
        ws_cmyk = wb.active
        ws_cmyk.title = "CMYK"

        # Iterando sobre cada pixel da imagem e adicionando os valores CMYK ao Excel
        for row, row_pixels in enumerate(img_rgb):
            for col, pixel in enumerate(row_pixels):
                # Convertendo os valores RGB para CMYK
                cmyk_values = rgb_to_cmyk(pixel)
                # Escrevendo os valores CMYK na célula correspondente
                ws_cmyk.cell(row=row + 1, column=col + 1).value = f"{cmyk_values[0]}, {cmyk_values[1]}, {cmyk_values[2]}, {cmyk_values[3]}"

        # Salvando o arquivo Excel
        wb.save(excel_path)
        print("Imagem convertida para Excel (CMYK) com sucesso!")

    def image_to_excelRGB(self, image_path, excel_path):
        # Leitura da imagem
        img = cv2.imread(image_path)
        # Conversão para RGB (OpenCV lê em BGR)
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # Criando um novo arquivo Excel
        wb = Workbook()
        ws_rgb = wb.active
        ws_rgb.title = "RGB"

        # Iterando sobre cada pixel da imagem e adicionando os valores RGB ao Excel
        for row, row_pixels in enumerate(img_rgb):
            for col, pixel in enumerate(row_pixels):
                # Obtendo os valores RGB do pixel
                r, g, b = pixel
                # Escrevendo os valores RGB na célula correspondente
                ws_rgb.cell(row=row + 1, column=col + 1).value = f"{r}, {g}, {b}"

        # Salvando o arquivo Excel
        wb.save(excel_path)
        print("Imagem convertida para Excel (RGB) com sucesso!")

    def image_to_matrix_gray(self, image_path):
        try:
            # Abre a imagem
            image = Image.open(image_path)
            # Converte a imagem em escala de cinza
            grayscale_image = image.convert('L')
            # Obtém as dimensões da imagem
            width, height = grayscale_image.size
            # Obtém os valores dos pixels da imagem usando numpy para eficiência
            pixel_values = np.array(grayscale_image.getdata(), dtype=np.uint8).reshape((height, width))
            return pixel_values
        except Exception as e:
            print("Erro ao converter imagem para matriz:", e)
            return None

    def save_matrix_to_excel_gray(self, matrix, excel_filename):
        try:
            # Cria um arquivo Excel
            wb = Workbook()
            ws = wb.active
            # Escreve os valores da matriz no arquivo Excel
            for row in matrix:
                ws.append(list(row))  # Convertendo numpy.ndarray para lista
            # Salva o Excel
            wb.save(excel_filename)
            #deu certo
            print("Imagem convertida para Excel (Gray) com sucesso!")
        except Exception as e:
            print("Erro fudeu tudo: ", e)

def select_image_convert(app):
        # Verifica se a imagem foi selecionada
        file_path = "IMG/winx.png"
        if file_path:
            image_gray = app.image_to_matrix_gray(file_path)
            app.image_to_excelRGB(file_path,"saidaRGB.xlsx")
            app.image_to_excel_cmyk(file_path,"saidaCMYK.xlsx")
            if image_gray is not None:
                app.save_matrix_to_excel_gray(image_gray, "saidaGray.xlsx")

            else:
                print("Não foi possível processar a imagem.")
        else:
            print("Nenhuma imagem foi selecionada.")



